from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

def write_to_excel(items, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Extracted'

    headers = ['#', 'Key', 'Value', 'Comments', 'RawContext']
    ws.append(headers)

    for idx, item in enumerate(items, start=1):
        ws.append([
            idx,
            item["key"],
            item["value"],
            item["comments"],
            item["raw_context"]
        ])

    for i in range(1, 6):
        ws.column_dimensions[get_column_letter(i)].width = 40

    wb.save(out_path)
